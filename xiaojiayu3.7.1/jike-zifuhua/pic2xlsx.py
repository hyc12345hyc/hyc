from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill,Fill

imageFileName = '2.bmp'  # 图片文件名
image = Image.open(imageFileName)  # 打开图片
wb = openpyxl.Workbook()  # 创建Excel
sheet = wb.create_sheet(imageFileName)  # 创建sheet
width,height = image.size  # 获取图片大小
image = image.resize((int(width * 0.5),int(height * 0.5)))  #调整图片大小
for w in range(image.width):
    for h in range(image.height):
        # 将每个像素的颜色填充到对应cell的背景色中
        rgba = image.getpixel((w, h))
        #也可以这样写
        # value = img.load()
        # rgba = value[w, h]
        # Python zfill() 方法返回指定长度的字符串，原字符串右对齐，前面填充0。
        colorHex = hex(rgba[0])[2:].zfill(2) + hex(rgba[1])[2:].zfill(2) + hex(rgba[2])[2:].zfill(2)
        fill = PatternFill(fill_type='solid', fgColor=colorHex)
        #fill = PatternFill(fill_type='solid', start_color=colorHex, end_color=colorHex)
        sheet.cell(row = h + 1, column = w + 1).fill = fill
wb.save(imageFileName + '.xlsx')  # 保存xlsx文件
