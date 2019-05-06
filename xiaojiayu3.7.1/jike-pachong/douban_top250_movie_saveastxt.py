import requests
import bs4
import openpyxl
a = []
head = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36'}
for x in range(0,2):
    res = requests.get("https://movie.douban.com/top250?start=%d" % (x*25),headers = head)
    soup = bs4.BeautifulSoup(res.text, "html.parser")
    title = soup.find_all("div", class_="hd")
    point = soup.find_all("span", property="v:average")
    staff = soup.find_all("p",class_ ="")
    for each,each1,each2 in zip(title,point,staff):
        a.append(each.a.span.text)
        #a.append('\n')
        a.append('评分：%s'%(each1.text))
        #a.append('\n')
        a.append(each2.text.replace(' ','').replace('\n',''))
        #a.append('\n')
        #a.append('\n')
# with open('douban.txt', 'w',encoding="utf-8") as f:
#     f.writelines(a)
def save_to_excel(result):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "电影名称"
    ws['B1'] = "评分"
    ws['C1'] = "资料"
    ws.append(result)

    wb.save("豆瓣TOP250电影.xlsx")
save_to_excel(a)


