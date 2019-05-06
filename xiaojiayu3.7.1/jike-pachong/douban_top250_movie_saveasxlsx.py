import requests
import bs4
import openpyxl
import re

movies = []
ranks = []
messages = []
years = []
head = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.75 Safari/537.36'}

for x in range(0,10):
    res = requests.get("https://movie.douban.com/top250?start=%d" % (x*25),headers = head)
    soup = bs4.BeautifulSoup(res.text, "html.parser")
    movie = soup.find_all("div", class_="hd")
    rank = soup.find_all("span", property="v:average")
    message = soup.find_all("p", class_="")
    for each in movie:
        movies.append(each.a.span.text)
    for each in rank:
        ranks.append(each.text)
    for each in message:
        messages.append(each.text.replace(' ', '').replace('\n', ''))
        years.append(re.sub("\D", "",each.text.replace(' ', '').replace('\n', '')))
    result = []
    length = len(movies)
    for i in range(length):
        result.append([movies[i], ranks[i], messages[i],years[i]])

def save_to_excel(result):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "电影名称"
    ws['B1'] = "评分"
    ws['C1'] = "资料"
    ws['D1'] = "年份"
    for each in result:
        ws.append(each)
    wb.save("豆瓣TOP250电影.xlsx")

save_to_excel(result)